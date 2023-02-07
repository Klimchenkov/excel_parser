import sys

import sqlite3 as sl
from openpyxl import load_workbook

class ReadExcel:

    TABLE_NAME = 'parsed_data'
    
    def __init__(self, path_to_file:str, parse:str):
        self.path_to_file = path_to_file
        self.con = sl.connect('excel_parser.db')
        self._create_db()
        if parse == 'true':
            self._parse_file()
        self._get_total()
        
    def _create_db(self):
        with self.con as con:
            sql = f"""
            CREATE TABLE IF NOT EXISTS {self.TABLE_NAME} (
                value INTEGER NOT NULL,
                date STRING,
                type STRING,
                company STRING,
                timing STRING
            );
            """
            con.execute(sql)

    def _parse_file(self):
        wb = load_workbook(self.path_to_file)
        sheet = wb.worksheets[0]
        for r in range(4, sheet.max_row + 1):
            for c in range(3, 11): 
                value = sheet.cell(row=r, column=c).value
                date = sheet.cell(row=3, column=c).value
                type = sheet.cell(row=2, column=c if c%2!=0 else c - 1).value
                company = sheet.cell(row=r, column=2).value
                timing = sheet.cell(row=1, column=3 if c in range(3,7) else 7).value
                tup = (value, date, type, company, timing)
                with self.con as con:
                    sql=f"""
                    INSERT INTO {self.TABLE_NAME} VALUES {tup};
                    """
                    con.execute(sql)

    def _get_total(self):
        
        with self.con as con:
            sql = 'SELECT date, timing, SUM(value) from parsed_data group by date, timing;'
            res = con.execute(sql)
            print('\nРезультат:\n')
            for i in res:
                print(f'Рассчётный {"" if i[1] == "fact" else "прогнозный "}тотал для {i[0]} - {i[2]}.')
            print('\n')
            
if __name__ == '__main__':
    ReadExcel(sys.argv[1], sys.argv[2])